import sqlite3
from sqlite3 import Error
import csv
import os
import openpyxl
import pandas as pd

# # # Connection to the database

def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
    return conn


# # # Main function

def main():
    # # Settings and open database file
    Folder = '/'.join(os.path.realpath(__file__)[:os.path.realpath(__file__).find(os. path. basename(__file__))].split(chr(92))[:-2])+'/'
    wb_obj = openpyxl.load_workbook(Folder + 'results_files/Input/' + 'inputs.xlsx', data_only = True)
    
    # list of the technologies of WAS sector for energy becnhmark
    sheet_obj = wb_obj['DO NOT CHANGE']
    list_tech_WAS = []
    tech_WAS = sheet_obj.cell(2 , 10)
    j=0
    while tech_WAS.value != None:
        list_tech_WAS.append(tech_WAS.value)
        j+=1
        tech_WAS = sheet_obj.cell(2 + j , 10)
        if j>48:
            break
    str_list_tech_WAS = "('" + "' , '".join([str(tech) for tech in list_tech_WAS]) + "')"

    # connection to database
    sheet_obj = wb_obj['DB_FILE']
    db_file = sheet_obj.cell(2 , 3).value
    conn = create_connection(Folder + 'data_files/' + db_file + '.sqlite')
    cur = conn.cursor()

    print('Starting and open database')


    # # Create new table with technologies' nomenclature in the database
    for sector_i in ["TRA" , "BDG" , "WAS" , "ENE"]:
        drop_table_if_exists = 'DROP TABLE IF EXISTS technology_nomenclature_' + sector_i
        cur.execute(drop_table_if_exists)
        if sector_i == "TRA":
            create_table = 'CREATE TABLE technology_nomenclature_TRA(Sub TEXT NOT NULL, Sector TEXT NOT NULL, Distance TEXT, Boundary TEXT, Purpose TEXT, Mode TEXT NOT NULL, Tech_type_1 TEXT, Tech_type_2 TEXT, Name TEXT, Vintage_Availability_year INTEGER, Efficiency TEXT, Energy_cat TEXT, Energy_bis TEXT, Label_tech TEXT, ID_TEMOA_TO TEXT NOT NULL, Description_TEMOA_TO TEXT NOT NULL, Subsector TEXT, Capacity_units TEXT, Output_units TEXT);'
            file = open(Folder + "results_files/Input/TRA/TRA_technology_nomenclature.csv")
            insert_records = 'INSERT INTO technology_nomenclature_TRA (Sub , Sector , Distance , Boundary , Purpose , Mode , Tech_type_1 , Tech_type_2 , Name , Vintage_Availability_year , Efficiency , Energy_cat , Energy_bis , Label_tech , ID_TEMOA_TO , Description_TEMOA_TO , Subsector, Capacity_Units, Output_Units) VALUES(? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? )'
            cur.execute(create_table)
            contents = csv.reader(file)
            cur.executemany(insert_records, contents)
            conn.commit()
        elif sector_i == "BDG":
            create_table = 'CREATE TABLE technology_nomenclature_BDG(Sub TEXT NOT NULL, Sector TEXT NOT NULL, Archetype TEXT, Vintage_Availability_year TEXT, Vintage_Building INTEGER, NewOld TEXT, Efficiency_reference TEXT, EndUse TEXT NOT NULL, Tech_type_1 TEXT, Tech_type_2 TEXT, Efficiency TEXT, Energy_cat TEXT, Energy_bis TEXT, Label_tech TEXT, ID_TEMOA_TO TEXT NOT NULL, Description_TEMOA_TO TEXT NOT NULL, Subsector TEXT, Capacity_units TEXT, Output_units TEXT);'
            file = open(Folder + "results_files/Input/BDG/BDG_technology_nomenclature.csv")
            insert_records = 'INSERT INTO technology_nomenclature_BDG (Sub , Sector , Archetype , Vintage_Availability_year , Vintage_Building , NewOld , Efficiency_reference , EndUse , Tech_type_1 , Tech_type_2 , Efficiency , Energy_cat , Energy_bis , Label_tech , ID_TEMOA_TO , Description_TEMOA_TO , Subsector, Capacity_Units, Output_Units) VALUES(? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? )'
            cur.execute(create_table)
            contents = csv.reader(file)
            cur.executemany(insert_records, contents)
            conn.commit()
        elif sector_i == "WAS":
            create_table = 'CREATE TABLE technology_nomenclature_WAS(Sub TEXT NOT NULL, Sector TEXT NOT NULL, EndUse TEXT, Treatment_Plant TEXT, Tech_type_1 TEXT, Tech_type_2 TEXT, Efficiency TEXT, Energy_cat TEXT, Vintage_Availability_year INTEGER , Label_tech TEXT, ID_TEMOA_TO TEXT NOT NULL, Description_TEMOA_TO TEXT NOT NULL, Subsector TEXT, Capacity_units TEXT, Output_units TEXT);'
            file = open(Folder + "results_files/Input/WAS/WAS_technology_nomenclature.csv")
            insert_records = 'INSERT INTO technology_nomenclature_WAS (Sub , Sector , EndUse , Treatment_Plant , Tech_type_1 , Tech_type_2 , Efficiency , Energy_cat , Vintage_Availability_year , Label_tech , ID_TEMOA_TO , Description_TEMOA_TO , Subsector, Capacity_Units, Output_Units) VALUES(? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? )'
            cur.execute(create_table)
            contents = csv.reader(file)
            cur.executemany(insert_records, contents)
            conn.commit()
        elif sector_i == "ENE":
            create_table = 'CREATE TABLE technology_nomenclature_ENE(Sub TEXT NOT NULL, Sector TEXT NOT NULL, EndUse TEXT, Energy_generated TEXT, Vintage_Availability_year INTEGER, Tech_cat TEXT, Tech_type_1 TEXT, Tech_type_2 TEXT, Tech_type_3 TEXT, Energy_cat TEXT, Label_tech TEXT, ID_TEMOA_TO TEXT NOT NULL, Description_TEMOA_TO TEXT NOT NULL, Subsector TEXT, Capacity_units TEXT, Output_units TEXT);'
            file = open(Folder + "results_files/Input/ENE/ENE_technology_nomenclature.csv")
            insert_records = 'INSERT INTO technology_nomenclature_ENE (Sub , Sector , EndUse , Energy_generated , Vintage_Availability_year , Tech_cat , Tech_type_1 , Tech_type_2 , Tech_type_3 , Energy_cat , Label_tech , ID_TEMOA_TO , Description_TEMOA_TO , Subsector, Capacity_Units, Output_Units) VALUES(? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? , ? )'
            cur.execute(create_table)
            contents = csv.reader(file)
            cur.executemany(insert_records, contents)
            conn.commit()

    print('Importing technologies nomenclature to database')


        # # Creation of the queries
    # creating a list with inputs queries sheets name
    list_sheets_name = wb_obj.sheetnames
    list_sheets_name_del = []
    del list_sheets_name[list_sheets_name.index('DO NOT CHANGE')]
    del list_sheets_name[list_sheets_name.index('DB_FILE')]
    for sheet_name in list_sheets_name: # delete architecture sheets that all begin by '--->'
        if sheet_name[:4] == "--->":
            list_sheets_name_del.append(sheet_name)
    for sheet_name in list_sheets_name_del:
        del list_sheets_name[list_sheets_name.index(sheet_name)]
    
    dict_db_table = {'Output_Emissions': 'SUM(Output_Emissions.emissions) AS Emisions ', 'Output_CapacityByPeriodAndTech': 'SUM(Output_CapacityByPeriodAndTech.capacity) AS Capacities ', 'Output_V_Capacity': 'SUM(Output_V_Capacity.capacity) AS Capacities ', 'Output_VFlow_In': 'SUM(Output_VFlow_In.vflow_in) AS VFlow_In ', 'Output_VFlow_Out': 'SUM(Output_VFlow_Out.vflow_out) AS VFlow_Out ' , 'Output_Duals': 'AVG(Output_Duals.dual) AS Marginal_Costs ', 'Output_Costs': 'SUM(Output_Costs.output_cost) AS Costs '}
    for sheet in list_sheets_name:
        sheet_obj = wb_obj[sheet]
        db_table = sheet_obj.cell(3 , 3).value
        description = sheet_obj.cell(2 , 3).value
        file_name = sheet_obj.cell(2 , 4).value
        # disaggregation_columns (SELECT, FROM and GROUP BY)
        list_disaggregation_columns = []
        disaggregation_columns = sheet_obj.cell(4 , 3).value
        i = 0
        while disaggregation_columns != None:
            if i >4:
                break
            i += 1
            list_disaggregation_columns.append(disaggregation_columns)
            disaggregation_columns = sheet_obj.cell(4 + i , 3).value
        list_disaggregation_columns_formula = []
        list_disaggregation_columns_formula_name = []
        disaggregation_columns_formula = sheet_obj.cell(9 , 3).value
        disaggregation_columns_formula_name = sheet_obj.cell(9 , 4).value
        if disaggregation_columns_formula != None:
            i = 0
            while disaggregation_columns_formula != None:
                if i >3:
                    break
                i += 1
                disaggregation_columns_formula += " AS '" + disaggregation_columns_formula_name +"'"
                list_disaggregation_columns_formula.append(disaggregation_columns_formula)
                list_disaggregation_columns_formula_name.append(disaggregation_columns_formula_name)
                disaggregation_columns_formula = sheet_obj.cell(9 + i , 3).value
                disaggregation_columns_formula_name = sheet_obj.cell(9 + i , 4).value
            if list_disaggregation_columns == []:
                str_disaggregation_columns = " " + " , ".join([str(col) for col in list_disaggregation_columns_formula])
                str_disaggregation_columns_and_formula_name = " " + " , ".join([str(name) for name in list_disaggregation_columns_formula_name])
            else:
                str_disaggregation_columns = " " + db_table + "." + (" , " + db_table + ".").join([str(col) for col in list_disaggregation_columns]) + " , " + " , ".join([str(col) for col in list_disaggregation_columns_formula])
                str_disaggregation_columns_and_formula_name = " " + db_table + "." + (" , " + db_table + ".").join([str(col) for col in list_disaggregation_columns]) + " , " + " , ".join([str(name) for name in list_disaggregation_columns_formula_name])
            str_GROUP_BY = "GROUP BY" + str_disaggregation_columns_and_formula_name
        else:
            str_disaggregation_columns = " " + db_table + "." + (" , " + db_table + ".").join([str(col) for col in list_disaggregation_columns])
            str_GROUP_BY = "GROUP BY" + str_disaggregation_columns


        list_nomenclature_columns = []
        list_nomenclature_columns_formula = []
        list_nomenclature_columns_formula_name = []
        nomenclature_table = sheet_obj.cell(13 , 3).value
        if nomenclature_table != None:
            nomenclature_columns = sheet_obj.cell(14 , 3).value
            i = 0
            while nomenclature_columns != None :
                if i > 14:
                    break
                i += 1
                list_nomenclature_columns.append(nomenclature_columns)
                nomenclature_columns = sheet_obj.cell(14 + i , 3).value
            nomenclature_columns_formula = sheet_obj.cell(29 , 3).value
            nomenclature_columns_formula_name = sheet_obj.cell(29 , 4).value
            if nomenclature_columns_formula != None:
                i = 0
                while nomenclature_columns_formula != None :
                    if i > 14:
                        break
                    i += 1
                    nomenclature_columns_formula += " AS '" + nomenclature_columns_formula_name + "'"
                    list_nomenclature_columns_formula.append(nomenclature_columns_formula)
                    list_nomenclature_columns_formula_name.append(nomenclature_columns_formula_name)
                    nomenclature_columns_formula = sheet_obj.cell(29 + i , 3).value
                    nomenclature_columns_formula_name = sheet_obj.cell(29 + i , 4).value
                str_nomenclature_columns = " " + nomenclature_table + "." + (" , " + nomenclature_table + "." ).join([str(col) for col in list_nomenclature_columns]) + " , " + " , ".join([str(col) for col in list_nomenclature_columns_formula])
                str_nomenclature_columns_and_formula_name = " " + nomenclature_table + "." + (" , " + nomenclature_table + "." ).join([str(col) for col in list_nomenclature_columns]) + " , " + " , ".join([str(name) for name in list_nomenclature_columns_formula_name])
                if str_GROUP_BY == "GROUP BY " + db_table + ".":
                    str_GROUP_BY = "GROUP BY "
                    str_GROUP_BY += str_nomenclature_columns_and_formula_name
                else:
                    str_GROUP_BY += " , " + str_nomenclature_columns_and_formula_name
            else:
                str_nomenclature_columns = " " + nomenclature_table + "." + (" , " + nomenclature_table + "." ).join([str(col) for col in list_nomenclature_columns])
                if str_GROUP_BY == "GROUP BY " + db_table + ".":
                    str_GROUP_BY = "GROUP BY "
                    str_GROUP_BY += str_nomenclature_columns
                else:
                    str_GROUP_BY += " , " + str_nomenclature_columns
            if str_disaggregation_columns == " " + db_table + ".":
                str_SELECT = "SELECT" + str_nomenclature_columns + " , " + dict_db_table[db_table]
                str_FROM = "FROM " + db_table + " , " + nomenclature_table + " "
            else:
                str_SELECT = "SELECT" + str_disaggregation_columns + " ," + str_nomenclature_columns + " , " + dict_db_table[db_table]
                str_FROM = "FROM " + db_table + " , " + nomenclature_table + " "
        else:
            str_SELECT = "SELECT" + str_disaggregation_columns + " , " + dict_db_table[db_table]
            str_FROM = "FROM " + db_table + " "


        # Conditions on columns (WHERE)
        list_conditions = []
        if nomenclature_table != None:
            list_conditions.append(db_table + ".tech = " + nomenclature_table + ".ID_TEMOA_TO ")
        conditions = sheet_obj.cell(35 , 3).value
        i=0
        while conditions != None:
            if i > 10:
                break
            i+=1
            if "str_list_tech_WAS" in conditions:
                conditions = conditions.replace("str_list_tech_WAS" , str_list_tech_WAS)
            list_conditions.append(conditions)
            conditions = sheet_obj.cell(35 + i , 3).value
        str_WHERE = "WHERE " + " AND ".join([str(condition) for condition in list_conditions]) + " "
        if str_WHERE == "WHERE  ":
            str_WHERE = ""

        # create query
        query = str_SELECT + str_FROM + str_WHERE + str_GROUP_BY
        # print('')
        # print(query)

        print('Creating query for ' + sheet + ' to file ' + file_name + '.csv')


            # # Execute query and filling CVS file
        dict_output_folder = {'E': 'Emissions', 'C': 'Capacities', 'N': 'Energy Supply', 'D': 'Costs' , 'A': 'End Use Energy Services'}
        if not os.path.exists(Folder + 'results_files/Output/' + db_file + '/' + dict_output_folder[sheet[0]]):
            os.makedirs(Folder + 'results_files/Output/' + db_file + '/' + dict_output_folder[sheet[0]])
        cvs_file = Folder + 'results_files/Output/' + db_file + '/' + dict_output_folder[sheet[0]] + '/' + file_name + '.csv'
        
        if(os.path.exists(cvs_file) and os.path.isfile(cvs_file)): # Delete previous CSV file
            os.remove(cvs_file)
            print('Previous .csv file deleted')
        cur.execute(query)
        
        columns = [column[0] for column in cur.description] # description of the columns
        data = [] # creation of an empty list used to fill the CSV file
        rows = cur.fetchall()
        for row in rows: # filling of the list as dictonnary object
            data.append(dict(zip(columns, row)))

        if nomenclature_table == None:
            df_dict = pd.read_excel(Folder + 'results_files/Input/' + 'Label_Dictionnary.xlsx', sheet_name = 'General', index_col = 0)
            label_dict = df_dict.to_dict()
        else:
            df_dict = pd.read_excel(Folder + 'results_files/Input/' + 'Label_Dictionnary.xlsx', sheet_name = nomenclature_table[-3:], index_col = 0)
            label_dict = df_dict.to_dict()            
        
        for i in range(len(data)):
            for key in data[i]:
                if data[i][key] in label_dict['Label'].keys():
                    data[i][key] = label_dict['Label'][data[i][key]] 

        with open(cvs_file, 'a', newline='') as new_file: # creation and filling of the CSV file
            first_line_writer = csv.writer(new_file)
            first_line_writer.writerow([description])
            fieldnames = columns # first line with description of the columns
            writer = csv.DictWriter(new_file, fieldnames = fieldnames)
            writer.writeheader()
            for line in data:
                writer.writerow(line)
        print('Exporting to cvs file done for ' + file_name)
        print('')
    
    conn.close()


main()


